# -*- coding: utf-8 -*-
import logging
import logging.handlers
import ConfigParser, os
from datetime import datetime
from pymongo import MongoClient
from alignment import Alignment
from alignment_set import AlignmentSet
from project import Project
from building import Building
from domain import Domain
from bson.objectid import ObjectId
import glob,io
import csv, re
import sys, getopt
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.tri as tri
from mpl_toolkits.mplot3d.axes3d import *
from scipy.interpolate import griddata
from collections import defaultdict
from shapely.geometry import LineString
from matplotlib.colors import LinearSegmentedColormap, ListedColormap
import osgeo.ogr as ogr
import osgeo.osr as osr
import osgeo.gdal as gdal
from utils import uz_laganathan,d_uz_dx_laganathan, d_ux_dx_laganathan,toFloat
from base import BaseStruct
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
# danzi.tn@20160322 plot sezioni a PK critiche
# create main logger 
"""
python plot_data_pk.py -a -p ../data -t ../data/pk_crit.txt
-p  folder di output
-t file di testo con una pk per riga
tipo
pk_crit.txt
---
2128218
2128378
2126198


"""
logger = logging.getLogger('smt_main')
logger.setLevel(logging.DEBUG)
# create a rotating file handler which logs even debug messages 
fh = logging.handlers.RotatingFileHandler('plto_data_pk.log',maxBytes=5000000, backupCount=5)
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s')
fh.setFormatter(formatter)
logger.addHandler(fh)
# reading config file
sCFGName = 'smt.cfg'
smtConfig = ConfigParser.RawConfigParser()
smtConfig.read(sCFGName)
# setup DB parameter
host = smtConfig.get('MONGODB','host')
database = smtConfig.get('MONGODB','database')
source_database = smtConfig.get('MONGODB','source_database')
username = smtConfig.get('MONGODB','username')
password = smtConfig.get('MONGODB','password')
# Colori da associare alle aree degli strati di riferimento
main_colors = {
                'MC':'#1f77b4',
                'MCA':'#ff7f0e', 
                'SO2':'#2ca02c',
                'MFL5':'#d62728',  
                'CG':'#9467bd',
                'MIG':'#8c564b',  
                'SB':'#e377c2', 
                'MA':'#7f7f7f',
                'SO4':'#bcbd22', 
                'MFL4':'#dbdb8d', 
                'SO':'#17becf', 
                'MFL3':'#9edae5',
                'R':'#7f7f7f',
                'AA':'#bcbd22',}

def mid_point(p1,p2):
    x = (p1[0] + p2[0])/2.
    y = (p1[1] + p2[1])/2.
    return x,y



def asse2p_p1(p1,p2,x):
    if p2[0]==p1[0]:
        return p1[1]
    if p2[1]==p1[1]:
        return None
    m = (p2[1]-p1[1])/(p2[0]-p1[0])
    m_asse = -1/m
    return m_asse*(x-p1[0]) + p1[1]
    

def pfromdistance_p1(p1,p2,d):
    if abs(p2[0]-p1[0])<=0.1:
        return (p1[0] - d,p1[1]) , (p1[0]+d,p1[1])
    if abs(p2[1]-p1[1])<=0.1:
        return (p1[0],p1[1]+d) , (p1[0],p1[1] - d)
    m = (p2[1]-p1[1])/(p2[0]-p1[0])
    mx = p1[0]
    a = 1.
    b = -2.*mx
    c = mx**2-d**2*m**2/(m**2+1)
    coeff = [a, b, c]
    res_x = np.roots(coeff)
    if isinstance(res_x[0], complex):
        print '%s=x is complex for %s %s %f' % (str(res_x[0]),p1,p2,d)
        print "p2[0]-p1[0] = %f" % (p2[0]-p1[0])
        print "p2[1]-p1[1] = %f" % (p2[1]-p1[1])
    res_y0 = asse2p_p1(p1,p2,res_x[0])
    res_y1 = asse2p_p1(p1,p2,res_x[1])
    return (res_x[0],res_y0),(res_x[1],res_y1)
                
                
def asse2p(p1,p2,x):
    mid = mid_point(p1,p2)
    if abs(p2[0]-p1[0])<=0.1:
        return mid[1]
    if abs(p2[1]-p1[1])<=0.1:
        return None
    m = (p2[1]-p1[1])/(p2[0]-p1[0])
    m_asse = -1/m
    return m_asse*(x-mid[0]) + mid[1]

def pfromdistance(p1,p2,d):
    mid = mid_point(p1,p2)
    if abs(p2[0]-p1[0])<=0.1:
        return (p1[0] - d,mid[1]) , (p1[0]+d,mid[1])
    if abs(p2[1]-p1[1])<=0.1:
        return (mid[0],p1[1]+d) , (mid[0],p1[1] - d)
    m = (p2[1]-p1[1])/(p2[0]-p1[0])
    mx = mid[0]
    a = 1.
    b = -2.*mx
    c = mx**2-d**2*m**2/(m**2+1)
    coeff = [a, b, c]
    res_x = np.roots(coeff)
    if isinstance(res_x[0], complex):
        print '%s=x is complex for %s %s %f' % (str(res_x[0]),p1,p2,d)
        print "p2[0]-p1[0] = %f" % (p2[0]-p1[0])
        print "p2[1]-p1[1] = %f" % (p2[1]-p1[1])
    res_y0 = asse2p(p1,p2,res_x[0])
    res_y1 = asse2p(p1,p2,res_x[1])
    return (res_x[0],res_y0),(res_x[1],res_y1)
        
                
def plot_line( ob, color="green"):
    parts = hasattr(ob, 'geoms') and ob or [ob]
    for part in parts:
        x, y = part.xy
        plt.plot(x, y,'o', color=color, zorder=1) 
        #plt.plot(x, y, color=color, linewidth=2, solid_capstyle='round', zorder=1)    


def plot_coords( x, y, color='#999999', zorder=1):
    plt.plot(x, y, 'o', color=color, zorder=zorder)


def outputFigure(sDiagramsFolderPath, sFilename, format="svg"):
    imagefname=os.path.join(sDiagramsFolderPath,sFilename)
    if os.path.exists(imagefname):
        os.remove(imagefname)
    plt.savefig(imagefname,format=format, bbox_inches='tight', pad_inches=0,  dpi=100)

def processSettlements(a_list):
    distanceIndex = defaultdict(list)
    for i, a_item in enumerate(a_list):
        for item in a_item["SETTLEMENTS"]:
            #if item["code"] > 0:
            distanceIndex[item["code"]].append(item["value"])
    distanceKeys = distanceIndex.keys()
    distanceKeys.sort()
    return distanceKeys, distanceIndex
    
def fillBetweenStrata(a_list):
    currentStrata = None
    x=[]
    y1=[]
    y2=[]
    facecolors=['orange','yellow']
    for i, a_item in enumerate(a_list):
        z_base = a_item['REFERENCE_STRATA']['POINTS']['base']['coordinates'][2]
        z_top = a_item['REFERENCE_STRATA']['POINTS']['top']['coordinates'][2]  
        code = a_item['REFERENCE_STRATA']['CODE']
        if not currentStrata:
            currentStrata = code        
        elif currentStrata == code:
            pass
        else:
            plt.fill_between(x, y1, y2,facecolor=main_colors[currentStrata], interpolate=True,label=currentStrata)
            x=[]
            y1=[]
            y2=[]
            currentStrata = code
        x.append(a_item['PK'])
        y1.append(z_base)
        y2.append(z_top)        

def plot_data_pk(bAuthenticate, sPath, sTxt):
    # connect to MongoDB
    client = MongoClient()
    db = client[database]
    ##################################################### GIS SETUP
    driver = ogr.GetDriverByName("ESRI Shapefile")
    # create the spatial reference, EPSG3949 RGF93 / CC49 - http://spatialreference.org/ref/epsg/3949/
    # xMin,yMin 1653513.80,8175914.21 : xMax,yMax 1659996.62,8177877.58
    srs = osr.SpatialReference()
    srs.ImportFromEPSG(3949)
    pk_crit = []
    with open(sTxt) as f:
        lines = f.readlines()
        for line in lines:
            pk_crit.append(toFloat(line))
    # DB authentication if required
    if bAuthenticate:
        bLoggedIn = db.authenticate(username,password,source=source_database)
    else:
        bLoggedIn = True
    if bLoggedIn:
        logger.info("Logged in")
        pd = db.Project.find_one({"project_code":"MFW001_0-010 Metro Paris-Ligne 15_T2A"})
        if pd:
            logger.info("Project %s found" % pd["project_name"])
            p = Project(db, pd)
            p.load()
            found_domains = Domain.find(db, {"project_id": p._id})
            for dom in found_domains:
                d = Domain(db,dom)
                d.load()
                # Example of query
                # asets = db.AlignmentSet.find({"$and":[{"domain_id": d._id},{"code": "main_01"}]})
                asets = db.AlignmentSet.find({"$and":[{"domain_id": d._id}]})
                for aset in asets:
                    a_set = AlignmentSet(db,aset)
                    a_set.load()
                    sCode = a_set.item["code"]
                    # {"$and":[{"alignment_set_id":"56f0fa102a13da2500a7250b"},{"PK":{"$in":[2126358,2126368,2126378]}}]}
                    criteria = {"$and":[{"alignment_set_id":a_set._id},{"PK":{"$in":pk_crit}}]} 
                    fields = {"PK":True,"COB":True,"BLOWUP":True, "PH":True, "DEM":True,"SETTLEMENT_MAX":True, "VOLUME_LOSS":True, "P_EPB":True,"SETTLEMENT_MAX_BASE":True, "VOLUME_LOSS_BASE":True,  "P_EPB_BASE":True, "STRATA":True, "SETTLEMENTS":True, "SECTIONS":True, "nu_tun":True, "beta_tun":True,  "BUILDINGS":True}
                    als = db.Alignment.find(criteria,fields).sort("PK", 1)
                    # eps0 = volume perso (adimensionale) = VOLUME_LOSS
                    # R = raggio di scavo in metri = SECTIONS.Excavation.Radius
                    # H = profondita' asse tunnel dalla superficie = z_dem - z_tun
                    # nu = coefficiente di poisson mediato dall'asse galleria alla superficie = nu_tun
                    # beta = 45° + coeff. di attrito/2 mediato dall'asse alla superficie. Per argille coeff attrito = 0 = beta_tun
                    # x distanza planimetrica ortogonale del punto di misura dall'asse -1.5*H, 1.5*H, 0.5
                    # z profondita' del punto di misura dalla superficie superficie
                    for al in als:
                        align = BaseStruct(al)
                        eps0 = align.VOLUME_LOSS
                        blowup = align.BLOWUP
                        eps0_base = align.VOLUME_LOSS_BASE
                        eps0_dict = {"base":eps0_base, "":eps0}
                        p_epb = align.P_EPB
                        p_epb_base = align.P_EPB_BASE
                        R = align.SECTIONS.Excavation.Radius
                        z_tun = align.PH.coordinates[2] + align.SECTIONS.Lining.Offset
                        z_dem = align.DEM.coordinates[2]
                        H = z_dem - z_tun
                        nu = align.nu_tun
                        beta = align.beta_tun
                        x = np.linspace(-2.5*H,2.5*H,51)
                        #x = np.arange(-1.5*H,1.5*H,1.0)
                        z = 0
                        y_uz = []
                        y_uz_base = []
                        y_duz = []
                        y_dux = []
                        y_duz_base = []
                        y_dux_base = []
                        pkxlabel = []
                        pkxticks = []
                        pkylabel_uz = []
                        pkyticks_uz = []
                        # excel strata sheet
                        wb = Workbook()
                        dest_filename = "%s_%d.xlsx" % (sCode,int(align.PK))
                        dest_filename = os.path.join(sPath,dest_filename)
                        ws1 = wb.active
                        ws1.title = "strata"
                        ws1.cell(column=1, row=1, value="CODE")
                        ws1.cell(column=2, row=1, value="Z_TOP")
                        ws1.cell(column=3, row=1, value="Z_BASE")
                        for row, s in enumerate(align.STRATA):
                            ws1.cell(column=1, row=row+2, value="%s" % s.CODE)
                            ws1.cell(column=2, row=row+2, value=s.POINTS.top.coordinates[2])
                            ws1.cell(column=3, row=row+2, value=s.POINTS.base.coordinates[2])
                        # excel buildings sheet
                        ws2 = wb.create_sheet(title="buildings")
                        ws2.cell(column=1, row=1, value="CODE")
                        ws2.cell(column=2, row=1, value="Dist MIN")
                        ws2.cell(column=3, row=1, value="Dist MAX")
                        ws2.cell(column=4, row=1, value="Z Fond")
                        ws2.cell(column=5, row=1, value="SC Lev")
                        ws2.cell(column=6, row=1, value="Damage Class Base")
                        ws2.cell(column=7, row=1, value="Damage Class")
                        ws2.cell(column=8, row=1, value="Vulnerability Class Base")
                        ws2.cell(column=9, row=1, value="Vulnerability Class")
                        ws2.cell(column=10, row=1, value="Settlement Max Base")
                        ws2.cell(column=11, row=1, value="Settlement Max")
                        ws2.cell(column=12, row=1, value="Tilt (Beta) Max Base")
                        ws2.cell(column=13, row=1, value="Tilt (Beta) Max")
                        ws2.cell(column=14, row=1, value="Horiz. Def. (Esp H) Max Base")
                        ws2.cell(column=15, row=1, value="Horiz. Def. (Esp H) Max")
                        ws2.cell(column=16, row=1, value="P Epb Base")
                        ws2.cell(column=17, row=1, value="P Epb")
                        ws2.cell(column=18, row=1, value="Blowup")
                        ws2.cell(column=19, row=1, value="Volume Loss Base")
                        ws2.cell(column=20, row=1, value="Volume Loss")
                        for row, b_align in enumerate(align.BUILDINGS):
                            b_dict = db.Building.find_one({"bldg_code":b_align.bldg_code})
                            b = BaseStruct(b_dict)
                            ws2.cell(column=1, row=row+2, value="%s" % b.bldg_code)
                            ws2.cell(column=2, row=row+2, value=b.d_min)
                            ws2.cell(column=3, row=row+2, value=b.d_max)
                            ws2.cell(column=4, row=row+2, value=b.depth_fondation)
                            ws2.cell(column=5, row=row+2, value=b.sc_lev)
                            ws2.cell(column=6, row=row+2, value=b.damage_class_base)
                            ws2.cell(column=7, row=row+2, value=b.damage_class)
                            ws2.cell(column=8, row=row+2, value=b.vulnerability_base)
                            ws2.cell(column=9, row=row+2, value=b.vulnerability)
                            ws2.cell(column=10, row=row+2, value=b.settlement_max_base)
                            ws2.cell(column=11, row=row+2, value=b.settlement_max)
                            ws2.cell(column=12, row=row+2, value=b.tilt_max_base)
                            ws2.cell(column=13, row=row+2, value=b.tilt_max)
                            ws2.cell(column=14, row=row+2, value=b.esp_h_max_base)
                            ws2.cell(column=15, row=row+2, value=b.esp_h_max)
                            ws2.cell(column=16, row=row+2, value=p_epb_base)
                            ws2.cell(column=17, row=row+2, value=p_epb)
                            ws2.cell(column=18, row=row+2, value=blowup)
                            ws2.cell(column=19, row=row+2, value=eps0_base)
                            ws2.cell(column=20, row=row+2, value=eps0)
                        wb.save(filename = dest_filename)
                                                
                        for i, xi in enumerate(x):
                            ret_uz = uz_laganathan(eps0, R, H, nu, beta, xi, z)
                            ret_uz_base = uz_laganathan(eps0_base, R, H, nu, beta, xi, z)
                            y_uz.append(-ret_uz)
                            y_uz_base.append(-ret_uz_base)
                            #if i%5 == 0:
                            if i == 25:
                                pkxticks.append(xi)
                                pkxlabel.append("%.2f m" % xi)
                                pkyticks_uz.append(-ret_uz)
                                pkylabel_uz.append("%.1f mm" % (ret_uz*1000.))
                            ret_uz = d_uz_dx_laganathan(eps0, R, H, nu, beta, xi, z)
                            ret_uz_base = d_uz_dx_laganathan(eps0_base, R, H, nu, beta, xi, z)
                            y_duz.append(ret_uz)      
                            y_duz_base.append(ret_uz_base)      
                            ret_uz = d_ux_dx_laganathan(eps0, R, H, nu, beta, xi, z) 
                            ret_uz_base = d_ux_dx_laganathan(eps0_base, R, H, nu, beta, xi, z)
                            y_dux.append(ret_uz)      
                            y_dux_base.append(ret_uz_base)      
                        # plot uz_laganathan
                        fig = plt.figure(figsize=(1200/100, 900/100), dpi=100)
                        plt.title(u"S[mm] PK %d" % (int(align.PK)))
                        ymin = min(y_uz)-(0.1*abs(min(y_uz)))
                        ymin = -0.03
                        plt.axis([min(x),max(x),ymin,0.0])
                        #plt.xticks(pkxticks, pkxlabel)
                        pkylabel_uz = [ "%.1f mm" % abs(yt*1000.) for yt in list(plt.yticks()[0])] + pkylabel_uz
                        plt.yticks(list(plt.yticks()[0]) + pkyticks_uz,pkylabel_uz)
                        plt.grid(True)
                        plt.plot(x,y_uz,"r-")
                        plt.plot(x,y_uz_base,"g--")
                        outputFigure(sPath, "S[mm]_uz_laganathan_%s_%d.png" % (sCode,int(align.PK)), format="png")                     
                        plt.close(fig)
                        # plot d_uz_dx_laganathan
                        # find min and max
                        pkylabel_dz = []
                        pkyticks_dz = []
                        pkxlabel = []
                        pkxticks = []
                        pkyticks_dz.append(min(y_duz))
                        pkylabel_dz.append(u"%.2f \u2030" % (min(y_duz)*1000.))
                        pkyticks_dz.append(max(y_duz))
                        pkylabel_dz.append(u"%.2f \u2030" % (max(y_duz)*1000.)) 
                        pkxticks.append(x[y_duz.index(min(y_duz))])
                        pkxticks.append(x[y_duz.index(max(y_duz))])
                        fig = plt.figure(figsize=(1200/100, 900/100), dpi=100)
                        plt.title(u"Beta[\u2030] PK %d" % (int(align.PK)))
                        plt.grid(True)
                        plt.plot(x,y_duz,"r-")
                        plt.plot(x,y_duz_base,"g--")
                        plt.axis([min(x),max(x),-10./1000,10./1000])
                        #plt.xticks(list(plt.xticks()[0]) + pkxticks)
                        pkylabel_dz = [ u"%.2f \u2030" % (yt*1000.) for yt in list(plt.yticks()[0])] + pkylabel_dz
                        plt.yticks(list(plt.yticks()[0]) + pkyticks_dz,pkylabel_dz)
                        outputFigure(sPath, u"Beta[\u2030]_d_uz_dx_laganathan_%s_%d.png" % (sCode,int(align.PK)), format="png")                         
                        plt.close(fig)
                        # plot d_uz_dx_laganathan
                        # find min and max
                        pkylabel_dx = []
                        pkyticks_dx = []
                        pkxlabel = []
                        pkxticks = []
                        pkyticks_dx.append(min(y_dux))
                        pkylabel_dx.append(u"%.2f \u2030" % (min(y_dux)*1000.))
                        pkyticks_dx.append(max(y_dux))
                        pkylabel_dx.append(u"%.2f \u2030" % (max(y_dux)*1000.)) 
                        pkxticks.append(x[y_dux.index(min(y_dux))])
                        pkxticks.append(x[y_dux.index(max(y_dux))])
                        fig = plt.figure(figsize=(1200/100, 900/100), dpi=100)
                        plt.title(u"EpsH[\u2030] PK %d" % (int(align.PK)))
                        plt.grid(True)
                        plt.plot(x,y_dux,"r-")
                        plt.plot(x,y_dux_base,"g--")
                        plt.axis([min(x),max(x),-1.5/1000,1.5/1000])
                        #plt.xticks(list(plt.xticks()[0]) + pkxticks)
                        pkylabel_dx = [ u"%.2f \u2030" % (yt*1000.) for yt in list(plt.yticks()[0])] + pkylabel_dx
                        plt.yticks(list(plt.yticks()[0]) + pkyticks_dx,pkylabel_dx)
                        outputFigure(sPath, u"Epsh[\u2030]_d_ux_dx_laganathan_%s_%d.png" % (sCode,int(align.PK)), format="png")                         
                        plt.close(fig)
                    logger.info("plot_data terminated!")
    else:
        logger.error("Authentication failed")
        
                
def main(argv):
    sPath = None
    sTxt = None
    bAuthenticate = False
    sSyntax = os.path.basename(__file__) +" -p <output folder path> -t <txt input for pk path> [-a for autentication -h for help]"
    try:
        opts, args = getopt.getopt(argv,"hap:t:",["path=","txt="])
    except getopt.GetoptError:
        print sSyntax
        sys.exit(1)
    if len(opts) < 2:
        print sSyntax
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print sSyntax
            sys.exit()
        elif opt == '-a':
            bAuthenticate = True
        elif opt in ("-p", "--path"):
            sPath = arg
            if not os.path.isdir(sPath):
                print sSyntax
                print "Directory %s does not exists!" % sPath
                sPath = None
                sys.exit(3)
        elif opt in ("-t", "--txt"):
            sTxt = arg
            if not os.path.isfile(sTxt):
                print sSyntax
                print "Directory %s does not exists!" % sTxt
                sTxt = None
                sys.exit(3)
    if sPath and sTxt:
        plot_data_pk(bAuthenticate, sPath, sTxt)
    
    
if __name__ == "__main__":
    main(sys.argv[1:])
